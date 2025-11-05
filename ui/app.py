import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import os
import threading
import csv
from datetime import datetime
import pandas as pd
import re
import shutil

from ui.theme import apply_theme
from ui.progress_window import ProgressWindow
from core.recorte_placas import recortar_placas
from core.marcar_rectangulos import marcar_rectangulos
from core.cortar_celdas import cortar_celdas
from core.detectar_radicula import calibrar_r_simple, analizar_radicula

class GerminIAApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("🌱 GerminIA")
        self.geometry("1280x800")
        self.minsize(1000, 700)
        self.state('zoomed') 
        apply_theme(self)
        self.modo = tk.StringVar(value="auto")
        self._current_df = None           # DataFrame que se muestra en la tabla (solo en Resultados)
        self._current_result_name = None  # Nombre de la carpeta de resultados actual
        self._build_ui()

    # ============================================================
    # INTERFAZ
    # ============================================================
    def _build_ui(self):
        sidebar = tk.Frame(self, bg="#fff", width=230, relief="flat", highlightbackground="#dee2e6", highlightthickness=1)
        sidebar.pack(side="left", fill="y")

        logo = tk.Label(sidebar, text="GerminIA", bg="#fff", fg="#0d6efd", font=("Segoe UI", 14, "bold"))
        logo.pack(pady=(20, 15))

        def add_btn(text, command, style="Accent.TButton"):
            b = ttk.Button(sidebar, text=text, command=command, style=style)
            b.pack(fill="x", padx=20, pady=5)
            return b

        add_btn("📁 Subir imágenes", self.subir_imagenes)
        add_btn("🕓 Ver pendientes", self.mostrar_pendientes)
        add_btn("⚙️ Procesar", self.ejecutar_procesamiento)
        ttk.Separator(sidebar, orient="horizontal").pack(fill="x", padx=20, pady=10)
        add_btn("🧩 Ver placas", lambda: self.mostrar_carpeta("placas"), "Neutral.TButton")
        #add_btn("🧫 Ver recortadas", lambda: self.mostrar_recortadas_resultados(), "Secondary.TButton")
        add_btn("📊 Ver resultados", lambda: self.mostrar_resultados_csv(), "Info.TButton")
        add_btn("🗑️ Eliminar resultados", self.eliminar_resultados, "Danger.TButton")

        # --- Botón Exportar a Excel (siempre al fondo) ---
        ttk.Separator(sidebar, orient="horizontal").pack(fill="x", padx=20, pady=(10, 6))
        self.btn_export_excel = ttk.Button(
            sidebar,
            text="📤 Exportar a Excel",
            style="Success.TButton",
            command=self.exportar_excel
        )
        self.btn_export_excel.pack(side="bottom", fill="x", padx=20, pady=(4, 14))
        self.btn_export_excel.state(["disabled"])  # deshabilitado salvo cuando hay resultados visibles


        # --- Contenido ---
        main = tk.Frame(self, bg="#f8f9fa")
        main.pack(side="right", fill="both", expand=True)

        title = tk.Label(main, text="Panel de Procesamiento", bg="#f8f9fa",
                         fg="#212529", font=("Segoe UI", 16, "bold"))
        title.pack(anchor="w", padx=20, pady=(15, 10))

        self.content_area = tk.Label(main, bg="#ffffff", relief="flat",
                                     font=("Segoe UI", 12), anchor="center")
        self.content_area.pack(expand=True, fill="both", padx=20, pady=10)

        self.status_text = tk.StringVar(value="Listo.")
        status_bar = ttk.Label(main, textvariable=self.status_text,
                               background="#e9ecef", anchor="w")
        status_bar.pack(fill="x", side="bottom")
        
        # Mostrar imágenes pendientes al iniciar
        self.after(200, self.mostrar_pendientes)


    # ============================================================
    # PROCESAMIENTO
    # ============================================================
    def _toggle_modo(self):
        messagebox.showinfo("Modo cambiado",
                            f"Modo actual: {'MANUAL' if self.modo.get() == 'manual' else 'AUTO'}")

    def _set_export_enabled(self, enabled: bool):
        if hasattr(self, "btn_export_excel"):
            if enabled:
                self.btn_export_excel.state(["!disabled"])
            else:
                self.btn_export_excel.state(["disabled"])

    
    # ============================================================
    # HELPERS
    # ============================================================ 
    @staticmethod    
    def _valid_recorte_folder(name: str) -> bool:
        """
        Acepta solo recortadas del tipo:
        recortes_placa_recortada..._<YYYYMMDD>_<YYYYMMDD>
        y además exige que ambas fechas sean iguales.
        """
        m = re.match(r"^recortes_placa_recortada.*?_(\d{8})_(\d{8})$", name)
        return bool(m and m.group(1) == m.group(2))

    # ============================================================
    # MOSTRAR IMAGENES SIN PROCESAR
    # ============================================================       
    def mostrar_pendientes(self):
        self._set_export_enabled(False)
        self.limpiar_vista()
        frame = tk.Frame(self.content_area.master, bg="#f8f9fa")
        frame.pack(fill="both", expand=True, padx=20, pady=15)

        ttk.Label(
            frame,
            text="📸 Imágenes pendientes de procesamiento",
            font=("Segoe UI", 16, "bold"),
            background="#f8f9fa"
        ).pack(pady=(0, 10))

        thumbs_frame = tk.Frame(frame, bg="#f8f9fa")
        thumbs_frame.pack(fill="both", expand=True)

        base = os.path.join("data", "germinacion", "data")
        base_originales = os.path.join(base, "originales")
        base_proc = os.path.join(base, "procesadas", "recortadas")
        base_res = os.path.join(base, "resultados")

        pendientes = []

        # --- 1️⃣ Buscar imágenes nuevas (no procesadas aún) ---
        for root, _, files in os.walk(base_originales):
            for f in files:
                if f.lower().endswith((".jpg", ".jpeg", ".png")):
                    nombre_base = os.path.splitext(f)[0]
                    # buscar si hay alguna carpeta de resultados que contenga ese nombre
                    ya_proc = any(nombre_base in c for c in os.listdir(base_res))
                    if not ya_proc:
                        pendientes.append((os.path.join(root, f), "nueva"))

        # --- 2️⃣ Buscar celdas recortadas sin su resultado ---
        for root, _, files in os.walk(base_proc):
            folder = os.path.basename(root)
            # ⛔️ Si es una carpeta recortadas con fechas distintas, la saltamos
            if folder.startswith("recortes_") and not self._valid_recorte_folder(folder):
                continue

            for f in files:
                if f.startswith("celda_") and f.lower().endswith(".jpg"):
                    path_celda = os.path.join(root, f)
                    nombre_carpeta = os.path.basename(root)
                    carpeta_resultados = os.path.join(base_res, nombre_carpeta)
                    path_res = os.path.join(carpeta_resultados, f"res_{f}")
                    if not os.path.exists(path_res):
                        pendientes.append((path_celda, "pendiente"))

        if not pendientes:
            ttk.Label(frame, text="✅ No hay imágenes pendientes.", background="#f8f9fa").pack(pady=40)
            return

        thumb_imgs = []

        for path_img, tipo in pendientes:
            img = Image.open(path_img)
            img.thumbnail((120, 120))
            tk_img = ImageTk.PhotoImage(img)

            def abrir_modal(ruta=path_img):
                top = tk.Toplevel(self)
                top.title(os.path.basename(ruta))
                top.configure(bg="#000")

                img = Image.open(ruta)
                iw, ih = img.size
                w, h = 1000, 750
                ratio = min(w / iw, (h - 140) / ih, 1)
                img = img.resize((int(iw * ratio), int(ih * ratio)))
                tk_big = ImageTk.PhotoImage(img)
                lbl = tk.Label(top, image=tk_big, bg="#000")
                lbl.image = tk_big
                lbl.pack(expand=True, pady=(10, 5))

                def calibrar():
                    csv_path = calibrar_r_simple(ruta)
                    if csv_path and os.path.exists(csv_path):
                        print(f"📄 CSV actualizado: {csv_path}")
                        self.mostrar_pendientes()  # refrescar vista

                #ttk.Button(
                    #top, text="🖋 Calibrar manual", style="Accent.TButton", command=calibrar
                #).pack(pady=10)

            color_borde = "#0d6efd" if tipo == "nueva" else "#ffc107"
            texto_tipo = "🆕" if tipo == "nueva" else "⏳"

            contenedor = tk.Frame(thumbs_frame, bg="#f8f9fa", bd=2, relief="flat")
            lbl_tipo = tk.Label(contenedor, text=texto_tipo, font=("Segoe UI", 10), bg="#f8f9fa", fg=color_borde)
            lbl_tipo.pack()
            btn = tk.Button(
                contenedor,
                image=tk_img,
                relief="flat",
                bg="#f8f9fa",
                activebackground="#f8f9fa",
                command=lambda r=path_img: abrir_modal(r)
            )
            btn.image = tk_img
            btn.pack()
            contenedor.pack(side="left", padx=6, pady=6)
            thumb_imgs.append(tk_img)

    def subir_imagenes(self):
        rutas = filedialog.askopenfilenames(
            title="Seleccionar imágenes",
            filetypes=[("Imágenes JPG/PNG", "*.jpg;*.png")]
        )
        if not rutas:
            return

        # Crear carpeta con la fecha actual
        fecha = datetime.now().strftime("%d%m%Y")
        destino_base = os.path.join("data", "germinacion", "data", "originales")
        destino_fecha = os.path.join(destino_base, f"originales_{fecha}")
        os.makedirs(destino_fecha, exist_ok=True)

        # Copiar imágenes dentro de la carpeta del día
        for ruta in rutas:
            nombre = os.path.basename(ruta)
            nuevo = os.path.join(destino_fecha, nombre)
            try:
                with open(ruta, "rb") as src, open(nuevo, "wb") as dst:
                    dst.write(src.read())
            except Exception as e:
                print(f"⚠️ Error copiando {nombre}: {e}")

        messagebox.showinfo(
            "Listo",
            f"{len(rutas)} imagen(es) añadidas en la carpeta:\n'originales_{fecha}'"
        )

    def ejecutar_procesamiento(self):
        modo = self.modo.get()
        logs = ProgressWindow(self)
        logs.show()

        def proceso():
            try:
                fecha = datetime.now().strftime("%d%m%Y")
                base = os.getcwd()

                originales_dir = os.path.join(base, "data", "germinacion", "data", "originales")
                procesadas_dir = os.path.join(base, "data", "germinacion", "data", "procesadas")
                resultados_dir = os.path.join(base, "data", "germinacion", "data", "resultados")

                os.makedirs(originales_dir, exist_ok=True)
                os.makedirs(procesadas_dir, exist_ok=True)
                os.makedirs(resultados_dir, exist_ok=True)

                subcarpetas = [f for f in os.listdir(originales_dir) if f.startswith("originales_")]
                if not subcarpetas:
                    print("⚠️ No hay carpetas con imágenes originales.")
                    messagebox.showinfo("Sin imágenes", "No hay carpetas con imágenes originales para procesar.")
                    return

                carpeta_reciente = max(subcarpetas, key=lambda c: os.path.getmtime(os.path.join(originales_dir, c)))
                carpeta_actual = os.path.join(originales_dir, carpeta_reciente)
                print(f"📂 Carpeta seleccionada: {carpeta_reciente}")

                originales = [
                    f for f in os.listdir(carpeta_actual)
                    if f.lower().endswith((".jpg", ".jpeg", ".png"))
                ]
                if not originales:
                    print("⚠️ No se encontraron imágenes para procesar.")
                    messagebox.showinfo("Sin imágenes", "No se encontraron imágenes nuevas para procesar.")
                    return

                nuevas = []
                for img in originales:
                    nombre_base = os.path.splitext(img)[0]
                    existe = any(nombre_base in carpeta for carpeta in os.listdir(resultados_dir))
                    if not existe:
                        nuevas.append(img)

                if not nuevas:
                    print("✅ Todas las imágenes ya fueron procesadas. Nada nuevo que hacer.")
                    messagebox.showinfo("Todo actualizado", "Todas las imágenes ya fueron procesadas.\nNo hay nada nuevo que hacer.")
                    return

                print(f"🧩 Imágenes nuevas detectadas: {len(nuevas)}")

                if modo == "manual":
                    print("🧠 Modo manual iniciado...")
                    carpeta = os.path.join(base, "data", "germinacion", "data", "procesadas", "recortadas")
                    calibrar_radicula_manual(carpeta)
                else:
                    print("🚀 Iniciando procesamiento automático...\n")
                    recortar_placas(base, fecha)
                    marcar_rectangulos(base, fecha)
                    cortar_celdas(base, fecha)
                    #analizar_radicula(base, fecha)
                    recortar_placas(base, fecha)
                    marcar_rectangulos(base, fecha)
                    cortar_celdas(base, fecha)

                    # 👉 Procesar SOLO los recortes del día (terminan en _{fecha}_{fecha})
                    recortadas_dir = os.path.join(base, "data", "germinacion", "data", "procesadas", "recortadas")
                    solo_hoy = sorted([
                        d for d in os.listdir(recortadas_dir)
                        if d.startswith("recortes_") and d.endswith(f"_{fecha}_{fecha}") and
                        os.path.isdir(os.path.join(recortadas_dir, d))
                    ])

                    if not solo_hoy:
                        print("ℹ️ No encontré recortes del día para analizar.")
                    else:
                        print("🔎 Analizando sólo:", solo_hoy)
                        analizar_radicula(base, fecha, solo_faltantes=True, solo_carpetas=solo_hoy)
                    print("✅ Proceso automático completado.")

            except Exception as e:
                print(f"❌ Error: {e}")
                messagebox.showerror("Error en procesamiento", f"Ocurrió un error:\n{e}")

        logs.run_in_thread(proceso)

    # ============================================================
    # VISUALIZACIÓN DE PLACAS
    # ============================================================
    def mostrar_carpeta(self, tipo):
        self._set_export_enabled(False)
        self.limpiar_vista()
        base = os.path.join("data", "germinacion", "data")
        carpeta = os.path.join(base, "procesadas", "placa_recortada")

        if not os.path.exists(carpeta):
            messagebox.showwarning("Sin datos", f"No se encontró la carpeta de placas.")
            return

        imagenes = [
            os.path.join(carpeta, f)
            for f in os.listdir(carpeta)
            if f.lower().endswith((".jpg", ".jpeg", ".png"))
        ]
        if not imagenes:
            messagebox.showinfo("Vacío", f"No hay imágenes en {carpeta}")
            return

        # === Guardamos las imágenes en la instancia ===
        self.imagenes = sorted(imagenes)
        self.indice_imagen = 0

        # === Frame principal ===
        frame = ttk.Frame(self.content_area.master)
        frame.pack(fill="both", expand=True, padx=20, pady=10)

        # === Título del nombre de la placa ===
        self.lbl_nombre_placa = ttk.Label(
            frame,
            text="",
            font=("Segoe UI", 12, "bold"),
            anchor="center"
        )
        self.lbl_nombre_placa.pack(pady=(5, 10))

        # === Imagen principal ===
        self.lbl_imagen = tk.Label(frame, bg="#f0f0f0")
        self.lbl_imagen.pack(expand=True)

        # === Navegación inferior ===
        botones_frame = ttk.Frame(frame)
        botones_frame.pack(pady=10)

        self.btn_anterior = ttk.Button(botones_frame, text="← Anterior", command=lambda: self.cambiar_imagen(-1))
        self.btn_anterior.grid(row=0, column=0, padx=5)

        self.btn_siguiente = ttk.Button(botones_frame, text="Siguiente →", command=lambda: self.cambiar_imagen(1))
        self.btn_siguiente.grid(row=0, column=1, padx=5)

        # Mostrar la primera imagen
        self.mostrar_imagen_actual()

        # Permitir cambiar con flechas del teclado
        self.bind_all("<Left>", lambda e: self.cambiar_imagen(-1))
        self.bind_all("<Right>", lambda e: self.cambiar_imagen(1))


    def mostrar_imagen_actual(self):
        """Muestra la imagen actual y el nombre de la placa arriba."""
        if not hasattr(self, "imagenes") or not self.imagenes:
            return

        ruta = self.imagenes[self.indice_imagen]
        nombre = os.path.basename(ruta)

        # Actualizar título
        self.lbl_nombre_placa.config(text=f"🧩 {nombre}")

        # Cargar imagen
        img = Image.open(ruta)
        img.thumbnail((950, 750))
        img_tk = ImageTk.PhotoImage(img)

        self.lbl_imagen.config(image=img_tk)
        self.lbl_imagen.image = img_tk


    def cambiar_imagen(self, paso):
        """Cambia de imagen con botones o flechas."""
        if not hasattr(self, "imagenes") or not self.imagenes:
            return

        self.indice_imagen = (self.indice_imagen + paso) % len(self.imagenes)
        self.mostrar_imagen_actual()

    # ============================================================
    # NUEVO: VER RESULTADOS (MUESTRA CSV)
    # ============================================================
    def mostrar_resultados_csv(self):
        base = os.path.join("data", "germinacion", "data", "resultados")
        if not os.path.exists(base):
            messagebox.showwarning("Sin datos", "No se encontró la carpeta de resultados.")
            return

        subcarpetas = [f for f in os.listdir(base) if os.path.isdir(os.path.join(base, f))]
        if not subcarpetas:
            messagebox.showinfo("Vacío", "No hay carpetas dentro de resultados.")
            return

        modal = tk.Toplevel(self)
        modal.title("Seleccionar carpeta con resultados")
        modal.geometry("420x250")
        modal.configure(bg="#f8f9fa")
        modal.resizable(False, False)
        modal.grab_set()

        # Centrar modal en pantalla
        modal.update_idletasks()
        x = (modal.winfo_screenwidth() - 420) // 2
        y = (modal.winfo_screenheight() - 250) // 2
        modal.geometry(f"+{x}+{y}")

        ttk.Label(modal, text="Seleccionar carpeta de resultados:",
                font=("Segoe UI", 10, "bold"), background="#f8f9fa").pack(pady=(15, 5))

        seleccion = tk.StringVar()
        combo = ttk.Combobox(modal, textvariable=seleccion, state="readonly", width=60)
        combo["values"] = subcarpetas
        combo.current(0)
        combo.pack(pady=10)

        def confirmar():
            modal.destroy()
            carpeta = os.path.join(base, seleccion.get())

            # Buscar cualquier archivo CSV
            csv_files = [f for f in os.listdir(carpeta) if f.lower().endswith(".csv")]
            if not csv_files:
                messagebox.showinfo("Sin archivo", f"No se encontró ningún archivo CSV en {carpeta}")
                return

            csv_path = os.path.join(carpeta, csv_files[0])
            self.mostrar_csv(csv_path, seleccion.get())

        ttk.Button(modal, text="Abrir", style="Accent.TButton", command=confirmar).pack(pady=(15, 10))
        ttk.Button(modal, text="Cancelar", command=modal.destroy).pack()

        modal.transient(self)
        modal.wait_window()


    # ============================================================
    # UTILITARIOS DE VISUALIZACIÓN DE IMÁGENES
    # ============================================================
    def mostrar_imagenes_directas(self, tipo, imagenes):
        if hasattr(self, "nav_frame") and self.nav_frame.winfo_exists():
            self.nav_frame.destroy()

        self.indice_img = 0
        self.imagenes = imagenes

        def mostrar_actual():
            ruta = self.imagenes[self.indice_img]
            img = Image.open(ruta)
            img.thumbnail((950, 650))
            img_tk = ImageTk.PhotoImage(img)
            self.content_area.configure(image=img_tk, text="")
            self.content_area.image = img_tk
            nombre = os.path.basename(ruta)
            self.status_text.set(f"{tipo.upper()} → {nombre} ({self.indice_img+1}/{len(self.imagenes)})")

        def siguiente():
            if self.indice_img < len(self.imagenes) - 1:
                self.indice_img += 1
                mostrar_actual()

        def anterior():
            if self.indice_img > 0:
                self.indice_img -= 1
                mostrar_actual()

        self.nav_frame = tk.Frame(self, bg="#e9ecef")
        self.nav_frame.pack(side="bottom", fill="x")
        ttk.Button(self.nav_frame, text="⟵ Anterior", command=anterior, style="Secondary.TButton").pack(side="left", padx=10, pady=5)
        ttk.Button(self.nav_frame, text="Siguiente ⟶", command=siguiente, style="Secondary.TButton").pack(side="left", padx=5, pady=5)

        mostrar_actual()
        
    def limpiar_vista(self):
        self._set_export_enabled(False)
        """Limpia la vista principal (imágenes o tablas)"""
        for widget in self.content_area.master.winfo_children():
            if widget not in [self.content_area]:
                widget.destroy()
        self.content_area.configure(image="", text="")
        if hasattr(self, "nav_frame") and self.nav_frame.winfo_exists():
            self.nav_frame.destroy()

    # ============================================================
    # MOSTRAR EL CSV
    # ============================================================
    def mostrar_csv(self, csv_path, carpeta_nombre):
        import pandas as pd
        from PIL import Image, ImageTk

        self.limpiar_vista()

        # === Frame principal ===
        frame = tk.Frame(self.content_area.master, bg="#f8f9fa")
        frame.pack(fill="both", expand=True, padx=20, pady=15)

        # === Título centrado ===
        title_frame = tk.Frame(frame, bg="#f8f9fa")
        title_frame.pack(fill="x", pady=(0, 10))
        ttk.Label(
            title_frame,
            text=f"📊 Resultados de {carpeta_nombre}",
            font=("Segoe UI", 16, "bold"),
            background="#f8f9fa",
            anchor="center"
        ).pack(anchor="center")
        
        # === Barra superior con comandos ===
        ayuda_frame = tk.Frame(frame, bg="#e9ecef", height=40)
        ayuda_frame.pack(fill="x", pady=(5, 10))

        label_ayuda = tk.Label(
            ayuda_frame,
            text="Controles:   R = borrar detección   |   Click = agregar detección   |   Z = deshacer   |   C = Codo   |   ENTER = guardar cambios   |   ESC = salir sin guardar",
            font=("Segoe UI", 9, "bold"),
            fg="#212529",
            bg="#e9ecef",
            pady=6
        )
        label_ayuda.pack(fill="x")

        # === Cargar CSV ===
        try:
            df = pd.read_csv(csv_path)

            # --- SANITIZAR NUMÉRICOS Y EVITAR NaN ---
            num_cols = ["Longitud_px", "Longitud_mm", "Longitud_cm", "Long_r", "Long_h"]
            for c in num_cols:
                if c in df.columns:
                    df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).round(2)

            # Guardamos para exportar y habilitamos botón
            self._current_df = df
            self._current_result_name = carpeta_nombre
            if hasattr(self, "btn_exportar"):
                self.btn_exportar.configure(state="normal")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo CSV:\n{e}")
            return

        self._current_df = df.copy()
        self._current_result_name = carpeta_nombre
        self._set_export_enabled(True)

        # === Mostrar miniaturas (grandes con nombre arriba) ===
        from PIL import Image, ImageTk
        import re

        carpeta = os.path.dirname(csv_path)

        def _num(f):
            m = re.search(r'celda_(\d+)', f)
            return int(m.group(1)) if m else 0

        # solo resultados (res_*.jpg), ordenados por número de celda
        img_files = sorted(
            [f for f in os.listdir(carpeta)
            if f.lower().endswith((".jpg", ".jpeg", ".png")) and f.startswith("res_")],
            key=_num
        )

        thumbs_outer = tk.Frame(frame, bg="#f8f9fa")
        thumbs_outer.pack(fill="x", pady=(5, 15))

        # --- Parámetros: ajustá acá si querés más grande ---
        THUMB_W, THUMB_H = 220, 220      # tamaño del “rectángulo”
        ROW_HEIGHT = THUMB_H + 80         # alto de la tira (mini + texto + márgenes)

        canvas = tk.Canvas(thumbs_outer, height=ROW_HEIGHT, bg="#f8f9fa", highlightthickness=0)
        inner = tk.Frame(canvas, bg="#f8f9fa")
        xscroll = ttk.Scrollbar(thumbs_outer, orient="horizontal", command=canvas.xview)
        canvas.configure(xscrollcommand=xscroll.set)

        canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.pack(fill="x")
        xscroll.pack(fill="x")

        thumb_imgs = []  # evitar GC

        def make_thumb(ruta: str) -> ImageTk.PhotoImage:
            img = Image.open(ruta).convert("RGB")
            iw, ih = img.size
            ratio = min(THUMB_W / iw, THUMB_H / ih)
            new_size = (max(1, int(iw * ratio)), max(1, int(ih * ratio)))
            img = img.resize(new_size, Image.LANCZOS)

            # letterbox para que TODAS queden exactamente del mismo tamaño
            bg = Image.new("RGB", (THUMB_W, THUMB_H), (233, 236, 239))  # #e9ecef para que contraste
            off_x = (THUMB_W - new_size[0]) // 2
            off_y = (THUMB_H - new_size[1]) // 2
            bg.paste(img, (off_x, off_y))
            return ImageTk.PhotoImage(bg)

        def mostrar_grande(ruta):
            top = tk.Toplevel(self)
            top.title(os.path.basename(ruta))
            top.configure(bg="#000")
            top.update_idletasks()
            w, h = 1000, 750
            sw, sh = top.winfo_screenwidth(), top.winfo_screenheight()
            top.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

            img = Image.open(ruta)
            iw, ih = img.size
            ratio = min(w / iw, (h - 140) / ih, 1)
            img = img.resize((int(iw * ratio), int(ih * ratio)), Image.LANCZOS)
            tk_img = ImageTk.PhotoImage(img)

            lbl = tk.Label(top, image=tk_img, bg="#000")
            lbl.image = tk_img
            lbl.pack(expand=True, pady=(10, 5))

            def calibrar_y_actualizar():
                csv_actualizado = calibrar_r_simple(ruta)
                if csv_actualizado and os.path.exists(csv_actualizado):
                    self.mostrar_csv(csv_actualizado, carpeta_nombre)

            ttk.Button(top, text="🖋 Calibrar manual", style="Accent.TButton",
                    command=calibrar_y_actualizar).pack(pady=(0, 5))

            ayuda = tk.Frame(top, bg="#1e1e1e")
            ayuda.pack(side="bottom", fill="x")
            tk.Label(
                ayuda,
                text="[R] Reiniciar  |  [Z] Deshacer  |  [C] Codo  |  [ENTER] Guardar  |  [ESC] Salir",
                font=("Segoe UI", 10), fg="#f0f0f0", bg="#1e1e1e", pady=6
            ).pack(fill="x")

            top.grab_set()

        for fname in img_files:
            ruta = os.path.join(carpeta, fname)
            tk_img = make_thumb(ruta)

            card = tk.Frame(inner, bg="#f8f9fa")
            # nombre ARRIBA (no sobre la imagen)
            tk.Label(card, text=fname, font=("Segoe UI", 9), bg="#f8f9fa").pack(pady=(0, 2))

            # usamos Label (no Button) y bind al click → menos problemas de render
            lbl = tk.Label(card, image=tk_img, bg="#f8f9fa", borderwidth=1, relief="solid",
                        highlightthickness=1, highlightbackground="#ced4da")
            lbl.image = tk_img               # retener referencia
            lbl.bind("<Button-1>", lambda e, r=ruta: mostrar_grande(r))
            lbl.pack()

            card.pack(side="left", padx=12, pady=10)
            thumb_imgs.append(tk_img)        # retener también acá

        # === Tabla ===
        table_frame = ttk.Frame(frame)
        table_frame.pack(fill="both", expand=True, pady=(10, 5))

        # Filtramos las columnas visibles
        cols_visibles = ["Celda", "Longitud_cm", "Long_r", "Long_h", "Estado"]

        tree = ttk.Treeview(
            table_frame,
            columns=cols_visibles,
            show="headings",
            height=20
        )

        # Encabezados
        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))

        for c in cols_visibles:
            tree.heading(c, text=c)
            ancho = 180 if c.lower() == "celda" else 140
            tree.column(c, anchor="center", stretch=True, width=ancho)

        # --- FORMATEADOR: 0 sin decimales, otros con 2 decimales ---
        def _fmt(v):
            # cadenas 'nan' -> '0'
            s = str(v).strip().lower()
            if s == "nan" or s == "":
                return "0"
            try:
                f = float(v)
                if abs(f) < 1e-12:
                    return "0"
                # si no es 0, mostramos 2 decimales
                return f"{f:.2f}"
            except:
                return str(v)

        # Insertar filas con color según estado
        for _, row in df.iterrows():
            values = [
                row.get("Celda", ""),
                _fmt(row.get("Longitud_cm", 0)),
                _fmt(row.get("Long_r", 0)),
                _fmt(row.get("Long_h", 0)),
                str(row.get("Estado", "")),
            ]
            estado = values[-1].upper()
            tag = "no_germinada" if "NO GERMINADA" in estado else "germinada"
            tree.insert("", "end", values=values, tags=(tag,))

        # Colores de filas
        tree.tag_configure("germinada", background="#d1f7c4")    # verde claro
        tree.tag_configure("no_germinada", background="#f7c4c4") # rojo claro

        # Scrollbar
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        tree.configure(yscroll=vsb.set)
        vsb.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True)

    # ============================================================
    # EXPORTAR A EXCEL
    # ============================================================
    def exportar_excel(self):

        if getattr(self, "_current_df", None) is None or self._current_df.empty:
            messagebox.showinfo("Exportar", "No hay resultados para exportar.")
            return

        # columnas que se exportan (lo mismo que ves en la tabla)
        cols = ["Celda", "Longitud_cm", "Long_r", "Long_h", "Estado"]
        df = self._current_df.copy()
         # --- EVITAR NaN EN LA EXPORTACIÓN ---
        for c in ["Longitud_cm", "Long_r", "Long_h"]:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).round(2)

        df = df[[c for c in cols if c in df.columns]]

        # diálogo guardar
        default_name = f"{(getattr(self, '_current_result_name', None) or 'resultados')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Guardar como",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel (*.xlsx)", "*.xlsx"), ("CSV (*.csv)", "*.csv")]
        )
        if not path:
            return

        # ¿CSV?
        if path.lower().endswith(".csv"):
            try:
                df.to_csv(path, index=False)
                messagebox.showinfo("Exportar", f"✅ CSV guardado:\n{path}")
            except Exception as e:
                messagebox.showerror("Exportar", f"No se pudo guardar el CSV:\n{e}")
            return

        # XLSX con formato en la columna Estado
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font
            from openpyxl.utils import get_column_letter

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Resultados")
                ws = writer.sheets["Resultados"]

                # Encabezados en negrita
                for cell in ws[1]:
                    cell.font = Font(bold=True)

                # Auto-anchos sencillos
                for idx, col in enumerate(df.columns, start=1):
                    ws.column_dimensions[get_column_letter(idx)].width = max(12, len(str(col)) + 2)

                # Colorear Estado
                if "Estado" in df.columns:
                    estado_col_idx = df.columns.get_loc("Estado") + 1
                    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # verde claro
                    fill_red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # rojo claro

                    for r in range(2, ws.max_row + 1):
                        cell = ws.cell(row=r, column=estado_col_idx)
                        v = str(cell.value).strip().upper() if cell.value is not None else ""
                        if "NO GERMINADA" in v:
                            cell.fill = fill_red
                        elif "GERMINADA" in v:
                            cell.fill = fill_green

            messagebox.showinfo("Exportar a Excel", f"✅ Archivo guardado:\n{path}")

        except ImportError:
            messagebox.showerror(
                "Exportar a Excel",
                "Falta el módulo 'openpyxl' para exportar con formato.\n"
                "Podés instalarlo con:\n\npip install openpyxl\n\n"
                "O elegí guardar como CSV en el diálogo."
            )
        except Exception as e:
            messagebox.showerror("Exportar a Excel", f"Ocurrió un error al exportar:\n{e}")

    # ============================================================
    # ELIMINAR IMAGENES
    # ============================================================

    def eliminar_resultados(self):
        """
        Vacía el contenido de:
        - data/germinacion/data/originales
        - data/germinacion/data/procesadas/placa_recortada
        - data/germinacion/data/procesadas/recortadas
        - data/germinacion/data/resultados
        Manteniendo SIEMPRE esas carpetas base.
        """
        base = os.path.join("data", "germinacion", "data")
        targets = [
            os.path.join(base, "originales"),
            os.path.join(base, "procesadas", "placa_recortada"),
            os.path.join(base, "procesadas", "recortadas"),
            os.path.join(base, "resultados"),
        ]

        # Confirmación
        msg = (
            "Se eliminará TODO el contenido (imágenes, subcarpetas y archivos)\n"
            "dentro de estas carpetas:\n\n"
            "• originales\n• procesadas/placa_recortada\n• procesadas/recortadas\n• resultados\n\n"
            "Las carpetas base permanecerán.\n\n¿Deseás continuar?"
        )
        if not messagebox.askyesno("Eliminar resultados", msg):
            return

        errores = []
        total_borrados = 0

        def _remove_readonly_and_retry(func, path, exc_info):
            # en Windows a veces hay archivos de solo lectura
            try:
                os.chmod(path, 0o700)
                func(path)
            except Exception as e:
                errores.append(f"{path}: {e}")

        try:
            for d in targets:
                if not os.path.isdir(d):
                    continue
                for name in os.listdir(d):
                    p = os.path.join(d, name)
                    try:
                        if os.path.isdir(p):
                            shutil.rmtree(p, onerror=_remove_readonly_and_retry)
                        else:
                            os.remove(p)
                        total_borrados += 1
                    except Exception as e:
                        errores.append(f"{p}: {e}")
        finally:
            # Refrescar UI y estados
            self._set_export_enabled(False)
            try:
                self.mostrar_pendientes()
            except Exception:
                pass

        if errores:
            messagebox.showwarning(
                "Eliminación parcial",
                "Se borró la mayoría del contenido, pero hubo errores en algunos elementos.\n\n"
                + "\n".join(errores[:10])  # mostramos hasta 10 para no saturar
                + ("\n\n(Se omitieron más errores…)" if len(errores) > 10 else "")
            )
        else:
            messagebox.showinfo("Eliminar resultados", f"✅ Limpieza finalizada. Elementos eliminados: {total_borrados}")

